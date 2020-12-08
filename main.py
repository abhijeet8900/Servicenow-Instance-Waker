from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time
import notification
import threading

MAX_TRY = 5


class wakeupInstanceThread(threading.Thread):
    def __init__(self, threadID, instance):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.instance = instance
        self.instanceID = instance['id']
        self.instanceUserName = instance['user_name']
        self.instancePassword = instance['password']
        self.instanceName = instance['name'] + '[' + self.instanceID + ']'

    def run(self):
        print(self.instanceName + ".....")
        self.__wakupInstance()

    def __wakupInstance(self):
        while self.instance['attempt'] < MAX_TRY:
            if self.instance['status'] != 'online':
                if self.instance['attempt'] > 0:
                    print('waiting for 5 min')
                    self.__externalWaitSeconds(300)
                self.instance['attempt'] += 1
                self.__setBrowserDriver()
                self.__loginToServiceNow()
                if self.__isInstanceOnline():
                    self.__markInstanceOnline()
                    self.webDriver.quit()
                    return True
                else:
                    self.__setInstanceStatusUnknown()
            else:
                if self.instance['attempt'] == MAX_TRY and self.instance['status'] == 'unknown':
                    self.__ErrorWithInstance()
                    self.webDriver.quit()
                    return False
                else:
                    self.__markInstanceOnline()
                    self.webDriver.quit()
                    return True

    def __implicitWait(self, sec):
        self.webDriver.implicitly_wait(sec)

    def __waitEleTobeSelecteable(self, id):
        return WebDriverWait(self.webDriver, 10).until(
            EC.element_to_be_clickable((By.ID, id))
        )

    def __externalWaitSeconds(self, sec):
        time.sleep(sec)

    def __isInstanceOnline(self):
        try:
            self.webDriver.get('https://'+self.instanceID+'.service-now.com/')
            self.webDriver.find_element_by_class_name(
                'instance-hibernating-page')
            return False
        except Exception as e:
            return True

    def __setBrowserDriver(self):
        options = Options()
        options.headless = True
        driver = webdriver.Firefox(
            options=options, executable_path='geckodriver.exe')
        driver.get("https://signon.service-now.com/ssologin.do?RelayState=%252Fapp%252Ftemplate_saml_2_0%252Fk317zlfESMUHAFZFXMVB%252Fsso%252Fsaml%253FRelayState%253Dhttps%25253A%25252F%25252Fdeveloper.servicenow.com%25252Fsaml_redirector.do%25253Fsysparm_nostack%25253Dtrue%252526sysparm_uri%25253D%2525252Fnav_to.do%2525253Furi%2525253D%252525252Fssologin.do%252525253FrelayState%252525253Dhttps%25252525253A%25252525252F%25252525252Fdeveloper.servicenow.com%25252525252Fdev.do&redirectUri=&email=")
        self.webDriver = driver

    def __loginToServiceNow(self):
        username_Input = self.webDriver.find_element_by_id('username')
        username_Input.send_keys(self.instanceUserName)
        self.webDriver.find_element_by_id('usernameSubmitButton').click()
        password_Input = self.__waitEleTobeSelecteable('password')
        password_Input.send_keys(self.instancePassword)
        self.__externalWaitSeconds(3)
        self.webDriver.find_element_by_id('submitButton').click()
        self.__implicitWait(5)
        self.__externalWaitSeconds(10)

    def __goToInstanceWakupPage(self):
        self.webDriver.get(
            'https://developer.servicenow.com/dev_app.do#!/instance?wu=true')
        self.__externalWaitSeconds(10)

    def __markInstanceOnline(self):
        print(self.instanceName+" : Online")
        self.instance['status'] = 'online'
        self.webDriver.quit()

    def __setInstanceStatusUnknown(self):
        print(self.instanceName + " : Offline Waking Instance...")
        self.instance['status'] = 'unknown'
        self.__goToInstanceWakupPage()
        self.webDriver.quit()

    def __ErrorWithInstance(self):
        print(self.instanceName + ' : Problem with instance ')
        print('https://'+self.instanceID+'.service-now.com/')
        self.webDriver.quit()


def getInstanceCredentials():
    wb = load_workbook(filename='instances.xlsx')
    ws = wb['Credentials']
    creds = []
    for row in ws.iter_rows(min_row=2):
        instance_name = row[0].value
        instance_user_name = row[1].value
        instance_password = row[2].value
        instance_id = row[3].value
        creds.append({
            "name": instance_name,
            "user_name": instance_user_name,
            "password": instance_password,
            "id": instance_id
        })
    return creds


def main():
    instanceCreds = getInstanceCredentials()
    for creds in instanceCreds:
        creds['status'] = 'unknown'
        creds['attempt'] = 0

    threadsBefore = threading.activeCount()
    for instance in instanceCreds:
        id = instance['id']
        t = wakeupInstanceThread(id, instance)
        t.start()

    while threadsBefore != threading.activeCount():
        time.sleep(1)

    Error = False
    for instance in instanceCreds:
        if instance['status'] != 'online':
            Error = True

    if Error:
        notification.fail_noise()
    else:
        notification.success_noise()


main()
